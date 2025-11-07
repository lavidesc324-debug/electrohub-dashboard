# -*- coding: utf-8 -*-
# Construye "Flujo_Caja_ElectroHub_v4_1a.xlsx" con:
# Parametros, Escenarios_v4, Flujo_Base, Optimizacion_Flujo, Indicadores,
# Dashboard (gráficos), KPIs_Escenarios (VAN/TIR/Payback/PI con método robusto),
# Resumen_Anual, Seguridad_H2, Instrucciones.
#
# Requiere: pip install openpyxl

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.workbook.defined_name import DefinedName

# ---------- utilería ----------
bold = Font(bold=True)
thin = Side(style="thin", color="999999")
border_bottom = Border(bottom=thin)
currency_fmt = '"$"#,##0;[RED]"$"#,##0'
pct_fmt = '0.00%'
int_fmt = '0'

def define_name(wb, name, ref):
    """Define un nombre para una referencia en el libro"""
    defn = DefinedName(name, attr_text=ref)
    wb.defined_names[name] = defn

def set_formula_by_label(ws, label, formula, numfmt=None):
    """Establece una fórmula en la celda a la derecha de una etiqueta"""
    for row in ws.rows:
        for cell in row:
            if cell.value == label:
                formula_cell = ws.cell(row=cell.row, column=cell.column + 1)
                formula_cell.value = formula
                if numfmt:
                    formula_cell.number_format = numfmt
                return formula_cell
    return None

# ---------- libro ----------
wb = Workbook()

# ----------------------- Parametros -----------------------
wsP = wb.active
wsP.title = "Parametros"
rows = [
    ["Descripción", "Valor", "Unidad / Nota"],
    ["Horizonte (años)", 15, "años"],
    ["Tasa de descuento (WACC)", 0.08, "porcentaje"],
    ["Inflación", 0.04, "porcentaje"],
    ["CAPEX total inicial (MXN)", 84_300_000, "Incluye terreno (60% año 0 / 40% año 1)"],
    ["Distribución CAPEX año 0", 0.60, "porcentaje"],
    ["Distribución CAPEX año 1", 0.40, "porcentaje"],
    ["Año reemplazo SAE", 10, "año"],
    ["Costo reemplazo SAE (MXN)", 4_000_000, ""],
    ["Valor residual terreno (MXN, año N)", 81_000_000, "año final"],
    ["Ingresos H2 Año operativo 1 (MXN)", 12_600_000, "+2%/año"],
    ["Ingresos O2 Año operativo 1 (MXN)", 6_720_000, "+2%/año"],
    ["Ingresos FV/red Año operativo 1 (MXN)", 2_740_000, "-0.7%/año (FV=1.2 MW)"],
    ["Ingresos carga EV Año operativo 1 (MXN)", 800_000, "+2%/año"],
    ["Crecimiento H2", 0.02, "porcentaje"],
    ["Crecimiento O2", 0.02, "porcentaje"],
    ["Crecimiento FV/red", -0.007, "porcentaje"],
    ["Crecimiento carga EV", 0.02, "porcentaje"],
    ["OPEX Base Año operativo 1 (MXN)", 2_600_000, "+3%/año"],
    ["Crecimiento OPEX Base", 0.03, "porcentaje"],
    ["OPEX Logística H2 Año op 1 (MXN)", 4_900_000, "+3%/año"],
    ["Crecimiento OPEX Logística H2", 0.03, "porcentaje"],
    ["OPEX Tratamiento O2 Año op 1 (MXN)", 2_200_000, "+3%/año"],
    ["Crecimiento OPEX Tratamiento O2", 0.03, "porcentaje"],
    ["Año inicio de operación", 2, "año (0 y 1 = inversión)"],
    ["Reserva liquidez (% OPEX Total)", 0.20, "optimización flujo de caja"],
    ["Tasa interés línea de crédito", 0.06, "porcentaje"],
]
for r, row in enumerate(rows, start=1):
    for c, val in enumerate(row, start=1):
        wsP.cell(r, c, val)

wsP["B2"].number_format = int_fmt
for addr in ["B3","B4","B6","B7","B15","B16","B17","B18","B20","B22","B24","B26","B27"]:
    wsP[addr].number_format = pct_fmt
for addr in ["B5","B9","B10","B11","B12","B13","B14","B19","B21","B23"]:
    wsP[addr].number_format = currency_fmt
wsP["B8"].number_format = int_fmt
for col in range(1, 4):
    wsP.column_dimensions[get_column_letter(col)].width = 40
for cell in wsP[1]:
    cell.font = bold

# nombres (para fórmulas entre hojas)
define_name(wb, 'HORIZONTE', "Parametros!$B$2")
define_name(wb, 'WACC', "Parametros!$B$3")
define_name(wb, 'INFLACION', "Parametros!$B$4")
define_name(wb, 'CAPEX', "Parametros!$B$5")
define_name(wb, 'CAPEX0', "Parametros!$B$6")
define_name(wb, 'CAPEX1', "Parametros!$B$7")
define_name(wb, 'ANIO_REP', "Parametros!$B$8")
define_name(wb, 'COSTO_REP', "Parametros!$B$9")
define_name(wb, 'VALOR_TERRENO', "Parametros!$B$10")
define_name(wb, 'ING_H2_1', "Parametros!$B$11")
define_name(wb, 'ING_O2_1', "Parametros!$B$12")
define_name(wb, 'ING_FV_1', "Parametros!$B$13")
define_name(wb, 'ING_EV_1', "Parametros!$B$14")
define_name(wb, 'G_H2', "Parametros!$B$15")
define_name(wb, 'G_O2', "Parametros!$B$16")
define_name(wb, 'G_FV', "Parametros!$B$17")
define_name(wb, 'G_EV', "Parametros!$B$18")
define_name(wb, 'OPEX_BASE_1', "Parametros!$B$19")
define_name(wb, 'G_OPEX_BASE', "Parametros!$B$20")
define_name(wb, 'OPEX_LOGH2_1', "Parametros!$B$21")
define_name(wb, 'G_OPEX_LOGH2', "Parametros!$B$22")
define_name(wb, 'OPEX_TRATO2_1', "Parametros!$B$23")
define_name(wb, 'G_OPEX_TRATO2', "Parametros!$B$24")
define_name(wb, 'ANIO_OP', "Parametros!$B$25")
define_name(wb, 'RESERVA_PCT', "Parametros!$B$26")
define_name(wb, 'TASA_CRED', "Parametros!$B$27")

# ----------------------- Escenarios_v4 -----------------------
wsE = wb.create_sheet("Escenarios_v4")
wsE.append(["Parámetro", "Valor Base (Año op 1)", "Unidad/Nota"])
base_rows = [
    ["ING_H2_BASE", 12600000, "MXN/año"],
    ["ING_O2_BASE",  6720000, "MXN/año"],
    ["ING_FV_BASE",  2740000, "MXN/año (FV=1.2 MW)"],
    ["ING_EV_BASE",   800000, "MXN/año"],
    ["OPEX_BASE_BASE",2600000, "MXN/año"],
    ["LOGH2_BASE",   4900000, "MXN/año"],
    ["TRATO2_BASE",  2200000, "MXN/año"],
]
for r in base_rows:
    wsE.append(r)
for c in wsE[1]:
    c.font = bold
    c.border = border_bottom
for r in range(2, 2 + len(base_rows)):
    wsE[f"B{r}"].number_format = currency_fmt
wsE.column_dimensions["A"].width = 24
wsE.column_dimensions["B"].width = 18
wsE.column_dimensions["C"].width = 38

wsE["E1"] = "Selector de escenario"; wsE["E1"].font = bold
wsE["E2"] = "Base"
wsE["G1"] = "Lista de escenarios"; wsE["G1"].font = bold
wsE["G2"] = "Optimista"; wsE["G3"] = "Base"; wsE["G4"] = "Conservador"
dv = DataValidation(type="list", formula1="=$G$2:$G$4")
wsE.add_data_validation(dv); dv.add(wsE["E2"])

wsE["A11"] = "Multiplicadores por escenario"; wsE["A11"].font = bold
headers = ["Escenario","f_H2","f_O2","f_FV","f_EV","f_OPEX_BASE","f_LOGH2","f_TRATO2"]
wsE.append(headers)
for i,h in enumerate(headers, start=1):
    wsE.cell(row=12, column=i).font = bold
    wsE.cell(row=12, column=i).border = border_bottom
mult = [
    ["Optimista",   170/150, 12/10, 1.10, 1.10, 0.95, 0.85, 0.90],
    ["Base",        1.00,    1.00,  1.00, 1.00, 1.00, 1.00, 1.00],
    ["Conservador", 130/150,  8/10, 0.90, 0.90, 1.05, 1.15, 1.10],
]
for row in mult:
    wsE.append(row)
for r in range(13, 16):
    for c in range(2, 9):
        wsE.cell(row=r, column=c).number_format = '0.00'

# Multiplicadores activos
wsE["J1"] = "Multiplicadores activos"; wsE["J1"].font = bold
labels = ["f_H2","f_O2","f_FV","f_EV","f_OPEX_BASE","f_LOGH2","f_TRATO2"]
for i,lbl in enumerate(labels, start=1):
    wsE.cell(row=1+i, column=10, value=lbl)
    wsE.cell(row=1+i, column=11, value=f"=INDEX($B$13:$H$15, MATCH($E$2,$A$13:$A$15,0), {i+1})")

# nombres
define_name(wb, 'ING_H2_BASE',     "Escenarios_v4!$B$2")
define_name(wb, 'ING_O2_BASE',     "Escenarios_v4!$B$3")
define_name(wb, 'ING_FV_BASE',     "Escenarios_v4!$B$4")
define_name(wb, 'ING_EV_BASE',     "Escenarios_v4!$B$5")
define_name(wb, 'OPEX_BASE_BASE',  "Escenarios_v4!$B$6")
define_name(wb, 'LOGH2_BASE',      "Escenarios_v4!$B$7")
define_name(wb, 'TRATO2_BASE',     "Escenarios_v4!$B$8")
define_name(wb, 'f_H2',            "Escenarios_v4!$K$2")
define_name(wb, 'f_O2',            "Escenarios_v4!$K$3")
define_name(wb, 'f_FV',            "Escenarios_v4!$K$4")
define_name(wb, 'f_EV',            "Escenarios_v4!$K$5")
define_name(wb, 'f_OPEX_BASE',     "Escenarios_v4!$K$6")
define_name(wb, 'f_LOGH2',         "Escenarios_v4!$K$7")
define_name(wb, 'f_TRATO2',        "Escenarios_v4!$K$8")

# enlazar Parametros a multiplicadores
set_formula_by_label(wsP, "Ingresos H2 Año operativo 1 (MXN)", "=ING_H2_BASE*f_H2", currency_fmt)
set_formula_by_label(wsP, "Ingresos O2 Año operativo 1 (MXN)", "=ING_O2_BASE*f_O2", currency_fmt)
set_formula_by_label(wsP, "Ingresos FV/red Año operativo 1 (MXN)", "=ING_FV_BASE*f_FV", currency_fmt)
set_formula_by_label(wsP, "Ingresos carga EV Año operativo 1 (MXN)", "=ING_EV_BASE*f_EV", currency_fmt)
set_formula_by_label(wsP, "OPEX Base Año operativo 1 (MXN)", "=OPEX_BASE_BASE*f_OPEX_BASE", currency_fmt)
set_formula_by_label(wsP, "OPEX Logística H2 Año op 1 (MXN)", "=LOGH2_BASE*f_LOGH2", currency_fmt)
set_formula_by_label(wsP, "OPEX Tratamiento O2 Año op 1 (MXN)", "=TRATO2_BASE*f_TRATO2", currency_fmt)

# ----------------------- Flujo_Base -----------------------
wsF = wb.create_sheet("Flujo_Base")
headers = [
    "Año","Ingresos H2","Ingresos O2","Ingresos FV/red","Ingresos EV",
    "Ingresos totales","OPEX Base","OPEX Log H2","OPEX Trat O2","OPEX Total",
    "CAPEX/Reemplazos","Valor residual",
    "Flujo neto","Factor descuento","Flujo descontado","VAN acumulado"
]
wsF.append(headers)
for c in wsF[1]:
    c.font = bold
    c.alignment = Alignment(horizontal="center")

H = wsP["B2"].value  # horizonte
for r in range(2, 2 + H + 1):
    wsF.cell(r, 1, r - 2)

# Año 0
wsF["B2"]="=0"; wsF["C2"]="=0"; wsF["D2"]="=0"; wsF["E2"]="=0"; wsF["F2"]="=0"
wsF["G2"]="=0"; wsF["H2"]="=0"; wsF["I2"]="=0"; wsF["J2"]="=0"
wsF["K2"]="=CAPEX*CAPEX0"; wsF["L2"]="=0"
wsF["M2"]="=F2-J2-K2+L2"
wsF["N2"]="=(1+WACC)^A2"; wsF["O2"]="=M2/N2"; wsF["P2"]="=O2"

# Años >=1
for r in range(3, 2 + H + 1):
    wsF[f"B{r}"]=f"=IF(A{r}<ANIO_OP,0,ING_H2_1*(1+G_H2)^(A{r}-ANIO_OP))"
    wsF[f"C{r}"]=f"=IF(A{r}<ANIO_OP,0,ING_O2_1*(1+G_O2)^(A{r}-ANIO_OP))"
    wsF[f"D{r}"]=f"=IF(A{r}<ANIO_OP,0,ING_FV_1*(1+G_FV)^(A{r}-ANIO_OP))"
    wsF[f"E{r}"]=f"=IF(A{r}<ANIO_OP,0,ING_EV_1*(1+G_EV)^(A{r}-ANIO_OP))"
    wsF[f"F{r}"]=f"=SUM(B{r}:E{r})"
    wsF[f"G{r}"]=f"=IF(A{r}<ANIO_OP,0,OPEX_BASE_1*(1+G_OPEX_BASE)^(A{r}-ANIO_OP))"
    wsF[f"H{r}"]=f"=IF(A{r}<ANIO_OP,0,OPEX_LOGH2_1*(1+G_OPEX_LOGH2)^(A{r}-ANIO_OP))"
    wsF[f"I{r}"]=f"=IF(A{r}<ANIO_OP,0,OPEX_TRATO2_1*(1+G_OPEX_TRATO2)^(A{r}-ANIO_OP))"
    wsF[f"J{r}"]=f"=G{r}+H{r}+I{r}"
    wsF[f"K{r}"]=f"=IF(A{r}=1,CAPEX*CAPEX1,IF(A{r}=ANIO_REP,COSTO_REP,0))"
    wsF[f"L{r}"]=f"=IF(A{r}=HORIZONTE,VALOR_TERRENO,0)"
    wsF[f"M{r}"]=f"=F{r}-J{r}-K{r}+L{r}"
    wsF[f"N{r}"]=f"=(1+WACC)^A{r}"
    wsF[f"O{r}"]=f"=M{r}/N{r}"
    wsF[f"P{r}"]=f"=P{r-1}+O{r}"

money_cols = "BCDEFGHIJKLMOP"
for col in money_cols:
    for r in range(2, 2+H+1):
        wsF[f"{col}{r}"].number_format = currency_fmt
for r in range(2, 2+H+1):
    wsF[f"N{r}"].number_format = '0.0000'
widths = [6,18,18,18,16,18,16,16,16,16,18,16,18,14,18,18]
for i,w in enumerate(widths, start=1):
    wsF.column_dimensions[get_column_letter(i)].width = w
wsF.freeze_panes = "A2"

# ----------------------- Optimizacion_Flujo -----------------------
wsO = wb.create_sheet("Optimizacion_Flujo")
headersO = ["Año","Flujo neto base","OPEX Total","Reserva requerida","Δ Reserva","Liquidez inicial",
            "Préstamo necesario","Interés deuda","Liquidez final","Deuda acumulada"]
wsO.append(headersO)
for c in wsO[1]:
    c.font = bold
    c.alignment = Alignment(horizontal="center")
for r in range(2, 2+H+1):
    wsO.cell(r, 1, r - 2)

wsO["B2"]="=Flujo_Base!M2"
wsO["C2"]="=Flujo_Base!J2"
wsO["D2"]="=C2*RESERVA_PCT"
wsO["E2"]="=D2"
wsO["F2"]="=0"
wsO["G2"]="=MAX(0, -(F2 + B2 - E2))"
wsO["H2"]="=0"
wsO["I2"]="=F2 + B2 - E2 + G2 - H2"
wsO["J2"]="=G2"

for r in range(3, 2+H+1):
    wsO[f"B{r}"]=f"=Flujo_Base!M{r}"
    wsO[f"C{r}"]=f"=Flujo_Base!J{r}"
    wsO[f"D{r}"]=f"=C{r}*RESERVA_PCT"
    wsO[f"E{r}"]=f"=D{r}-D{r-1}"
    wsO[f"F{r}"]=f"=I{r-1}"
    wsO[f"G{r}"]=f"=MAX(0, -(F{r} + B{r} - E{r}))"
    wsO[f"H{r}"]=f"=J{r-1}*TASA_CRED"
    wsO[f"I{r}"]=f"=F{r} + B{r} - E{r} + G{r} - H{r}"
    wsO[f"J{r}"]=f"=J{r-1} + G{r} + H{r}"

for col in "BCDEFGHIJ":
    for r in range(2, 2+H+1):
        wsO[f"{col}{r}"].number_format = currency_fmt
wsO.column_dimensions["A"].width = 6
for col in "BCDEFGHIJ":
    wsO.column_dimensions[col].width = 18
wsO.freeze_panes = "A2"

# ----------------------- Indicadores -----------------------
wsI = wb.create_sheet("Indicadores")
wsI.append(["Año","VAN acumulado (Flujo_Base)","TIR acumulada"])
for c in wsI[1]:
    c.font = bold
    c.alignment = Alignment(horizontal="center")
for r in range(2, 2+H+1):
    wsI.cell(r, 1, r - 2)
    wsI[f"B{r}"]=f"=Flujo_Base!P{r}"
    wsI[f"C{r}"]=f"=IF(A{r}=0,NA(),IRR(Flujo_Base!M$2:INDEX(Flujo_Base!M:M, A{r}+2)))"
wsI.column_dimensions["A"].width = 6
wsI.column_dimensions["B"].width = 24
wsI.column_dimensions["C"].width = 18

# ----------------------- Dashboard (gráficos) -----------------------
wsD = wb.create_sheet("Dashboard")
wsD["A1"]="Resumen Ejecutivo – ElectroHub (v4.1a con logística H2/O2, FV=1.2 MW y escenarios)"
wsD["A1"].font = Font(bold=True, size=13)
wsD.column_dimensions["A"].width = 100

# Flujo Neto
ch1 = LineChart(); ch1.title="Flujo Neto Anual"; ch1.y_axis.title="MXN"; ch1.x_axis.title="Año"
ch1.add_data(Reference(wsF, min_col=13, min_row=1, max_row=2+H), titles_from_data=True)
ch1.set_categories(Reference(wsF, min_col=1, min_row=2, max_row=2+H))
wsD.add_chart(ch1, "A3")

# Liquidez
ch2 = LineChart(); ch2.title="Liquidez Final por Año"; ch2.y_axis.title="MXN"; ch2.x_axis.title="Año"
ch2.add_data(Reference(wsO, min_col=9, min_row=1, max_row=2+H), titles_from_data=True)
ch2.set_categories(Reference(wsO, min_col=1, min_row=2, max_row=2+H))
wsD.add_chart(ch2, "A20")

# VAN acumulado
ch3 = LineChart(); ch3.title="VAN Acumulado"; ch3.y_axis.title="MXN"; ch3.x_axis.title="Año"
ch3.add_data(Reference(wsI, min_col=2, min_row=1, max_row=2+H), titles_from_data=True)
ch3.set_categories(Reference(wsI, min_col=1, min_row=2, max_row=2+H))
wsD.add_chart(ch3, "M3")

# OPEX Total
ch4 = LineChart(); ch4.title="OPEX Total"; ch4.y_axis.title="MXN"; ch4.x_axis.title="Año"
ch4.add_data(Reference(wsF, min_col=10, min_row=1, max_row=2+H), titles_from_data=True)
ch4.set_categories(Reference(wsF, min_col=1, min_row=2, max_row=2+H))
wsD.add_chart(ch4, "M20")

# ----------------------- KPIs_Escenarios -----------------------
wsK = wb.create_sheet("KPIs_Escenarios")
wsK.append(["Escenario","VAN (MXN)","TIR","Payback (años)","PI"])
for c in wsK[1]:
    c.font = bold
    c.border = border_bottom

wsK.append(["Año","Flujo_Optimista","Flujo_Base","Flujo_Conservador"])
for i in range(4):
    wsK.cell(2, 1+i).font = bold
for r in range(3, 3+16):
    wsK.cell(r, 1, r - 3)

# Multiplicadores por escenario (helpers en columnas G..I)
wsK["G2"]="Optimista"; wsK["G2"].font=bold
for i in range(7):
    wsK.cell(3+i,7, f"=INDEX(Escenarios_v4!$B$13:$H$15, MATCH($G$2,Escenarios_v4!$A$13:$A$15,0), {i+2})")
wsK["H2"]="Base"; wsK["H2"].font=bold
for i in range(7):
    wsK.cell(3+i,8, 1)  # Base = 1
wsK["I2"]="Conservador"; wsK["I2"].font=bold
for i in range(7):
    wsK.cell(3+i,9, f"=INDEX(Escenarios_v4!$B$13:$H$15, MATCH($I$2,Escenarios_v4!$A$13:$A$15,0), {i+2})")

def set_flow_col(col_letter, base_col_letter):
    # col_letter: B/C/D; base_col_letter: G/H/I
    for r in range(3, 3+16):
        wsK[f"{col_letter}{r}"] = (
            f"=IF(A{r}<ANIO_OP,0,"
            f"(Escenarios_v4!$B$2*{base_col_letter}$3*(1+G_H2)^(A{r}-ANIO_OP)"
            f"+Escenarios_v4!$B$3*{base_col_letter}$4*(1+G_O2)^(A{r}-ANIO_OP)"
            f"+Escenarios_v4!$B$4*{base_col_letter}$5*(1+G_FV)^(A{r}-ANIO_OP)"
            f"+Escenarios_v4!$B$5*{base_col_letter}$6*(1+G_EV)^(A{r}-ANIO_OP))"
            f"-(Escenarios_v4!$B$6*{base_col_letter}$7*(1+G_OPEX_BASE)^(A{r}-ANIO_OP)"
            f"+Escenarios_v4!$B$7*{base_col_letter}$8*(1+G_OPEX_LOGH2)^(A{r}-ANIO_OP)"
            f"+Escenarios_v4!$B$8*{base_col_letter}$9*(1+G_OPEX_TRATO2)^(A{r}-ANIO_OP)))"
            f"+IF(A{r}=0,CAPEX*CAPEX0,IF(A{r}=1,CAPEX*CAPEX1,IF(A{r}=ANIO_REP,COSTO_REP,0)))*(-1)"
            f"+IF(A{r}=HORIZONTE,VALOR_TERRENO,0)"
            f")"
        )

set_flow_col("B","G")  # Optimista
set_flow_col("C","H")  # Base
set_flow_col("D","I")  # Conservador

# KPIs (incluye año 0 en la serie para IRR y VAN)
wsK["A20"] = "KPIs por Escenario"; wsK["A20"].font = bold
wsK["A21"] = "Optimista"; wsK["A22"] = "Base"; wsK["A23"] = "Conservador"

wsK["B21"] = "=NPV(WACC, B4:B19) + B3"
wsK["B22"] = "=NPV(WACC, C4:C19) + C3"
wsK["B23"] = "=NPV(WACC, D4:D19) + D3"
wsK["C21"] = "=IRR(B3:B19)"; wsK["C22"] = "=IRR(C3:C19)"; wsK["C23"] = "=IRR(D3:D19)"

# Payback robusto con flujos descontados acumulados (helpers E..K)
wsK["E19"] = "Helper (desc & cum)"; wsK["E19"].font = bold
wsK["E20"] = "Año"; wsK["F20"] = "B_desc"; wsK["G20"] = "B_cum"; wsK["H20"] = "C_desc"; wsK["I20"] = "C_cum"; wsK["J20"] = "D_desc"; wsK["K20"] = "D_cum"
for c in "EFGHIJK":
    wsK[f"{c}20"].font = bold
for r in range(21, 21+16):
    wsK.cell(r,5, r-21)  # Año
    wsK[f"F{r}"] = f"=B{r-17}/(1+WACC)^E{r}"
    wsK[f"G{r}"] = f"=IF(E{r}=0, F{r}, G{r-1}+F{r})"
    wsK[f"H{r}"] = f"=C{r-17}/(1+WACC)^E{r}"
    wsK[f"I{r}"] = f"=IF(E{r}=0, H{r}, I{r-1}+H{r})"
    wsK[f"J{r}"] = f"=D{r-17}/(1+WACC)^E{r}"
    wsK[f"K{r}"] = f"=IF(E{r}=0, J{r}, K{r-1}+J{r})"

wsK["D21"] = "=MATCH(TRUE,INDEX(G21:G36>=0,0),0)-1"
wsK["D22"] = "=MATCH(TRUE,INDEX(I21:I36>=0,0),0)-1"
wsK["D23"] = "=MATCH(TRUE,INDEX(K21:K36>=0,0),0)-1"

for r in [21,22,23]:
    wsK[f"B{r}"].number_format = currency_fmt
    wsK[f"C{r}"].number_format = '0.00%'
    wsK[f"D{r}"].number_format = '0.0'
wsK["E20"] = "PI"; wsK["E20"].font = bold
wsK["E21"] = "=NPV(WACC,B4:B19)/ABS(B3)"
wsK["E22"] = "=NPV(WACC,C4:C19)/ABS(C3)"
wsK["E23"] = "=NPV(WACC,D4:D19)/ABS(D3)"

# Gráfico VAN por escenario
chart = BarChart()
chart.title = "VAN por escenario"
chart.add_data(Reference(wsK, min_col=2, min_row=20, max_row=23), titles_from_data=True)
chart.set_categories(Reference(wsK, min_col=1, min_row=21, max_row=23))
wsK.add_chart(chart, "M3")

# ----------------------- Resumen_Anual -----------------------
wsR = wb.create_sheet("Resumen_Anual")
headers = ["Año","H₂ (MXN)","O₂ (MXN)","FV/red (MXN)","EV (MXN)",
           "Ingresos totales (MXN)","OPEX Base","OPEX Log H₂","OPEX Trat O₂","OPEX Total",
           "CAPEX/Reemplazos","Valor residual","Flujo neto","VAN acumulado",
           "Liquidez final","Deuda acumulada"]
wsR.append(headers)
for c in wsR[1]:
    c.font = bold
    c.alignment = Alignment(horizontal="center")
    c.border = border_bottom
for r in range(2, 2+H+1):
    yr = r - 2
    wsR.cell(r,1, yr)
    wsR[f"B{r}"] = f"=Flujo_Base!B{r}"
    wsR[f"C{r}"] = f"=Flujo_Base!C{r}"
    wsR[f"D{r}"] = f"=Flujo_Base!D{r}"
    wsR[f"E{r}"] = f"=Flujo_Base!E{r}"
    wsR[f"F{r}"] = f"=Flujo_Base!F{r}"
    wsR[f"G{r}"] = f"=Flujo_Base!G{r}"
    wsR[f"H{r}"] = f"=Flujo_Base!H{r}"
    wsR[f"I{r}"] = f"=Flujo_Base!I{r}"
    wsR[f"J{r}"] = f"=Flujo_Base!J{r}"
    wsR[f"K{r}"] = f"=Flujo_Base!K{r}"
    wsR[f"L{r}"] = f"=Flujo_Base!L{r}"
    wsR[f"M{r}"] = f"=Flujo_Base!M{r}"
    wsR[f"N{r}"] = f"=Flujo_Base!P{r}"
    wsR[f"O{r}"] = f"=Optimizacion_Flujo!I{r}"
    wsR[f"P{r}"] = f"=Optimizacion_Flujo!J{r}"
for col in range(2, 16+1):
    wsR.column_dimensions[get_column_letter(col)].width = 18
    for r in range(2, 2+H+1):
        wsR[f"{get_column_letter(col)}{r}"].number_format = currency_fmt
wsR.column_dimensions["A"].width = 6
wsR.freeze_panes = "A2"

# ----------------------- Seguridad_H2 -----------------------
wsS = wb.create_sheet("Seguridad_H2")
wsS["A1"] = "Parámetros de seguridad y flashback – Hidrógeno (H₂)"; wsS["A1"].font = bold
wsS["A3"] = "Velocidad laminar de combustión (Su)"; wsS["B3"] = 2.86; wsS["C3"] = "m/s (≈286 cm/s)"
wsS["A5"] = "Normas recomendadas"; wsS["A5"].font = bold
wsS["A6"] = "ISO 5175-1"; wsS["B6"] = "Arrestadores de retroceso (estaciones/puntos de uso)"
wsS["A7"] = "ISO 16852"; wsS["B7"] = "Apagallamas en líneas/venteos (deflagración/detonación)"
wsS["A8"] = "NFPA 2 / CGA G-5.5"; wsS["B8"] = "Distancias/criterios de seguridad y venteo"
wsS.column_dimensions["A"].width = 36; wsS.column_dimensions["B"].width = 90; wsS.column_dimensions["C"].width = 34

# ----------------------- Instrucciones -----------------------
wsIns = wb.create_sheet("Instrucciones")
wsIns["A1"]="Cómo usar la plantilla v4.1a"; wsIns["A1"].font = bold
steps=[
"1) Elige escenario en Escenarios_v4 (Optimista/Base/Conservador).",
"2) Parametros toma automáticamente los valores iniciales × multiplicadores del escenario.",
"3) Flujo_Base y Optimizacion_Flujo se recalculan por fórmula; no edites manualmente celdas calculadas.",
"4) KPIs_Escenarios compara VAN/TIR/Payback/PI de los 3 escenarios en paralelo (con payback robusto).",
"5) Resumen_Anual lista por año ingresos, OPEX por componente, flujo neto, VAN, liquidez y deuda.",
"6) Seguridad_H2 contiene el dato de flashback (Su=2.86 m/s) y normas de referencia (ISO/NFPA/CGA)."
]
for i,t in enumerate(steps, start=3):
    wsIns.cell(i,1,t)
wsIns.column_dimensions["A"].width = 120

# Guardar
salida = "Flujo_Caja_ElectroHub_v4_1a.xlsx"
wb.save(salida)
print(f"OK -> {salida}")
